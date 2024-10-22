import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# 准备数据
data = {
    '学生姓名': ['张三', '李四', '王五', '赵六'],
    '高数': [85, 90, 78, 92],
    '英语': [90, 85, 92, 88],
    '计算机': [88, 82, 85, 90],
    '思想品德': [95, 88, 90, 85]
}

df = pd.DataFrame(data)

# 计算总分和平均分
df['总分'] = df[['高数', '英语', '计算机', '思想品德']].sum(axis=1)
df['平均分'] = df[['高数', '英语', '计算机', '思想品德']].mean(axis=1)

# 按总分降序排列
df_sorted = df.sort_values(by='总分', ascending=False)
# 保存到Excel
df_sorted.to_excel('学生成绩表.xlsx', index=False)

# 使用openpyxl设置格式
wb = load_workbook('学生成绩表.xlsx')
ws = wb.active

# 设置所有单元格居中对齐
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center')

# 设置第一行字体和底纹
for cell in ws[1]:
    cell.font = Font(size=16, name='宋体')
    cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# 设置其他行字体
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        if cell.column_letter in ['A']:  # 学生姓名
            cell.font = Font(size=14, name='宋体')
        else:
            cell.font = Font(size=14, name='Times New Roman')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# 添加表格框线
for row in ws.iter_rows():
    for cell in row:
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# 保存
wb.save('学生成绩表格式化.xlsx')