import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference

# 准备数据
data = {
    '姓名': ['张三', '李四', '王五', '赵六', '孙七', '周八', '吴九', '郑十', '王十一'],
    '基本工资': [3000, 3200, 3100, 3300, 3500, 3400, 3600, 3700, 3800],
    '奖金': [500, 600, 550, 580, 620, 600, 650, 700, 750],
    '补贴': [200, 180, 220, 210, 240, 230, 250, 270, 300],
    '缺勤扣款': [50, 40, 30, 20, 10, 0, 5, 10, 20],
    '扣所得税': [150, 160, 140, 130, 120, 110, 100, 90, 80]
}

# 创建DataFrame
df = pd.DataFrame(data)

# 计算应发工资和实发工资
df['应发工资'] = df[['基本工资', '奖金', '补贴']].sum(axis=1)
df['实发工资'] = df['应发工资'] - df['缺勤扣款'] - df['扣所得税']

# 自动填充编号
df['编号'] = ['A0' + str(i+1).zfill(2) for i in range(len(df))]

# 保存到Excel
df.to_excel('工资发放明细表.xlsx', index=False)

# 使用openpyxl设置格式
wb = load_workbook('工资发放明细表.xlsx')
ws = wb.active

# 合并单元格
ws.merge_cells('C1')
ws.merge_cells('D1')

# 设置字体和对齐方式
for row in ws.iter_rows(min_row=1, max_row=1):
    for cell in row:
        cell.font = Font(size=20, name='宋体', bold=True)
        cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal='center')

for row in ws.iter_rows(min_row=2, max_row=2):
    for cell in row:
        cell.font = Font(size=16, name='宋体')
        cell.alignment = Alignment(horizontal='center')

for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
    for cell in row:
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        if cell.column_letter == 'A':  # 编号
            cell.font = Font(size=14, name='宋体')
        elif cell.column_letter == 'B':  # 姓名
            cell.font = Font(size=14, name='宋体')
        elif cell.column_letter == 'C':  # 单位
            cell.alignment = Alignment(horizontal='left')
            cell.font = Font(size=14, name='宋体')
        elif cell.column_letter == 'D':  # 日期
            cell.alignment = Alignment(horizontal='right')
            cell.font = Font(size=14, name='宋体')
        else:
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(size=14, name='Times New Roman')

# 插入柱状图
chart = BarChart()
data = Reference(ws, min_col=7, min_row=4, max_row=12)
cats = Reference(ws, min_col=2, min_row=4, max_row=12)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.title = "实发工资柱状图"
chart.style = 10
chart.x_axis.title = "姓名"
chart.y_axis.title = "实发工资"

# 设置图表格式
chart.x_axis.title_font = Font(size=11, name='黑体')
chart.y_axis.title_font = Font(size=11, name='黑体')
chart.title_font = Font(size=18, name='宋体', bold=True)

ws.add_chart(chart, "E4")

# 保存
wb.save('242440507靳飞宇Excel作业.xlsx')